"""
Table Store Service — manages SQLite storage for tabular data (CSV/Excel)
and provides Ollama-powered table description and SQL generation.

Ingestion
---------
1. Parse uploaded CSV / Excel file into headers + rows (per sheet).
2. Call Ollama to generate a natural-language description and column types.
3. Create a typed SQLite table and bulk-insert all rows.
4. Register the table in an internal ``_table_registry``.

Query
-----
1. Accept a natural-language question + list of table IDs.
2. Fetch schemas from the registry.
3. Call Ollama to generate a ``SELECT`` query.
4. Execute the query (read-only, with retry on error).
5. Return structured results with citation metadata.
"""

from __future__ import annotations

import csv
import io
import json
import logging
import re
import sqlite3
import uuid
from pathlib import Path
from typing import Any, Optional

import httpx

from src.config import settings

logger = logging.getLogger(__name__)

# ── SQLite type mapping ──────────────────────────────────────

SQLITE_TYPE_MAP = {
    "TEXT": "TEXT",
    "INTEGER": "INTEGER",
    "REAL": "REAL",
    "DATE": "TEXT",
    "BOOLEAN": "INTEGER",
    "NUMERIC": "NUMERIC",
}

# ── Name sanitisation helpers ────────────────────────────────


def _sanitize_identifier(name: str, prefix: str = "t") -> str:
    """Create a safe SQLite identifier from arbitrary text."""
    name = re.sub(r"[^a-zA-Z0-9_]", "_", str(name))
    name = re.sub(r"_+", "_", name).strip("_").lower()
    if not name or name[0].isdigit():
        name = f"{prefix}_{name}"
    return name[:64]


# ── Service ──────────────────────────────────────────────────


class TableStoreService:
    """Manages SQLite storage for tabular data and Ollama integration."""

    def __init__(self) -> None:
        self.db_path: str = settings.sqlite_db_path
        self.ollama_url: str = settings.ollama_url
        self.ollama_model: str = settings.ollama_table_model
        self.text_to_sql_model: str = settings.ollama_text_to_sql_model
        self._http_client: Optional[httpx.AsyncClient] = None
        self._ensure_db()

    # ── Lazy HTTP client ─────────────────────────────────────

    @property
    def http_client(self) -> httpx.AsyncClient:
        if self._http_client is None:
            self._http_client = httpx.AsyncClient(timeout=300.0)
        return self._http_client

    # ── SQLite helpers ───────────────────────────────────────

    def _get_conn(self, readonly: bool = False) -> sqlite3.Connection:
        uri = f"file:{self.db_path}"
        if readonly:
            uri += "?mode=ro"
        conn = sqlite3.connect(uri, uri=True)
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA foreign_keys=ON")
        conn.execute("PRAGMA busy_timeout=5000")
        conn.row_factory = sqlite3.Row
        return conn

    def _ensure_db(self) -> None:
        Path(self.db_path).parent.mkdir(parents=True, exist_ok=True)
        conn = self._get_conn()
        try:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS _table_registry (
                    table_id    TEXT PRIMARY KEY,
                    table_name  TEXT NOT NULL UNIQUE,
                    source_path TEXT NOT NULL,
                    file_hash   TEXT NOT NULL,
                    sheet_name  TEXT,
                    description TEXT,
                    columns_json TEXT,
                    row_count   INTEGER DEFAULT 0,
                    created_at  TEXT DEFAULT (datetime('now'))
                )
                """
            )
            conn.commit()
        finally:
            conn.close()
        logger.info("SQLite table-store initialised at %s", self.db_path)

    # ── File parsing ─────────────────────────────────────────

    def parse_table_file(self, file_bytes: bytes, filename: str) -> list[dict]:
        """Parse a CSV / Excel file.

        Returns a list of dicts, one per sheet::

            [{"sheet_name": str | None, "headers": [...], "rows": [[...], ...]}]
        """
        ext = Path(filename).suffix.lower()
        if ext == ".csv":
            return self._parse_csv(file_bytes)
        if ext == ".tsv":
            return self._parse_csv(file_bytes, delimiter="\t")
        if ext in (".xlsx", ".xls", ".ods"):
            return self._parse_excel(file_bytes)
        raise ValueError(f"Unsupported table file format: {ext}")

    @staticmethod
    def _parse_csv(
        file_bytes: bytes, delimiter: str = ","
    ) -> list[dict]:
        text = file_bytes.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        all_rows = list(reader)
        if not all_rows:
            return []
        headers = all_rows[0]
        return [{"sheet_name": None, "headers": headers, "rows": all_rows[1:]}]

    @staticmethod
    def _parse_excel(file_bytes: bytes) -> list[dict]:
        import openpyxl

        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=True, data_only=True
        )
        results: list[dict] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [list(row) for row in ws.iter_rows(values_only=True)]
            if len(rows) < 2:
                continue  # need at least header + 1 data row
            headers = [
                str(h) if h is not None else f"col_{i}"
                for i, h in enumerate(rows[0])
            ]
            results.append(
                {"sheet_name": sheet_name, "headers": headers, "rows": rows[1:]}
            )
        wb.close()
        return results

    # ── Ollama: table description ────────────────────────────

    async def describe_table_with_ollama(
        self,
        headers: list[str],
        sample_rows: list[list[Any]],
        file_name: str,
        sheet_name: str | None = None,
    ) -> dict:
        """Call Ollama to generate a description and column-type mapping."""
        sheet_info = f", sheet '{sheet_name}'" if sheet_name else ""

        sample_csv = ", ".join(headers) + "\n"
        for row in sample_rows[:3]:
            sample_csv += ", ".join(str(v) if v is not None else "" for v in row) + "\n"

        prompt = (
            f'You are a data analyst. Analyze this table from file "{file_name}"{sheet_info}.\n\n'
            f"Headers: {', '.join(headers)}\n"
            f"Sample rows (first {min(len(sample_rows), 3)}):\n{sample_csv}\n"
            "Respond with ONLY this JSON (no markdown, no explanation):\n"
            "{\n"
            '  "description": "2-3 sentence description of what this table contains",\n'
            '  "columns": [\n'
            '    {"name": "exact_header_name", "type": "TEXT|INTEGER|REAL|DATE|BOOLEAN", '
            '"description": "brief column purpose"}\n'
            "  ]\n"
            "}\n\n"
            "Rules:\n"
            '- "name" must exactly match the provided headers.\n'
            "- Include ALL columns.\n"
            '- "type" must be one of: TEXT, INTEGER, REAL, DATE, BOOLEAN.\n'
        )

        resp = await self.http_client.post(
            f"{self.ollama_url}/api/generate",
            json={
                "model": self.ollama_model,
                "prompt": prompt,
                "stream": False,
                "format": "json",
                "options": {"temperature": 0.1, "num_predict": 1000, "num_ctx": 4096},
                "keep_alive": "10m",
            },
        )
        resp.raise_for_status()

        raw = resp.json().get("response", "{}")
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            logger.warning("Ollama response not valid JSON — using TEXT fallback")
            return {
                "description": f"Table from {file_name}{sheet_info}",
                "columns": [
                    {"name": h, "type": "TEXT", "description": ""} for h in headers
                ],
            }

    # ── Table creation / population ──────────────────────────

    def create_and_populate_table(
        self,
        table_id: str,
        headers: list[str],
        column_types: list[dict],
        rows: list[list[Any]],
        source_path: str,
        file_hash: str,
        sheet_name: str | None,
        description: str,
    ) -> str:
        """Create a SQLite table, insert data, and register it.

        Returns the sanitised table name.
        """
        base = Path(source_path).stem
        if sheet_name:
            base = f"{base}_{sheet_name}"
        table_name = _sanitize_identifier(base, prefix="t")

        # Ensure uniqueness by appending short hash if name already used
        conn = self._get_conn()
        try:
            existing = conn.execute(
                "SELECT 1 FROM _table_registry WHERE table_name = ? AND table_id != ?",
                (table_name, table_id),
            ).fetchone()
            if existing:
                table_name = f"{table_name}_{table_id[:8]}"

            # Build column type map from Ollama response
            type_map: dict[str, str] = {}
            for col in column_types:
                type_map[col["name"]] = SQLITE_TYPE_MAP.get(
                    col.get("type", "TEXT"), "TEXT"
                )

            # Build CREATE TABLE statement with sanitised column names
            col_defs: list[str] = []
            sanitised_headers: list[str] = []
            for h in headers:
                san = _sanitize_identifier(h, prefix="col")
                sqlite_type = type_map.get(h, "TEXT")
                col_defs.append(f'"{san}" {sqlite_type}')
                sanitised_headers.append(san)

            conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
            conn.execute(f'CREATE TABLE "{table_name}" ({", ".join(col_defs)})')

            # Insert rows
            placeholders = ", ".join(["?"] * len(headers))
            insert_sql = f'INSERT INTO "{table_name}" VALUES ({placeholders})'

            typed_rows: list[list[Any]] = []
            for row in rows:
                typed = []
                for i, val in enumerate(row):
                    expected = type_map.get(headers[i], "TEXT") if i < len(headers) else "TEXT"
                    typed.append(self._coerce_value(val, expected))
                # Pad or truncate to match header count
                while len(typed) < len(headers):
                    typed.append(None)
                typed_rows.append(typed[: len(headers)])

            conn.executemany(insert_sql, typed_rows)

            # Register
            conn.execute(
                """
                INSERT OR REPLACE INTO _table_registry
                    (table_id, table_name, source_path, file_hash, sheet_name,
                     description, columns_json, row_count)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    table_id,
                    table_name,
                    source_path,
                    file_hash,
                    sheet_name,
                    description,
                    json.dumps(column_types),
                    len(rows),
                ),
            )
            conn.commit()
            logger.info(
                "Created table '%s' with %d rows (%d cols)",
                table_name,
                len(rows),
                len(headers),
            )
            return table_name
        finally:
            conn.close()

    # ── Deletion ─────────────────────────────────────────────

    def delete_tables_by_source(self, source_path: str) -> list[str]:
        """Delete all tables originating from *source_path*.

        Returns a list of deleted table IDs.
        """
        conn = self._get_conn()
        try:
            rows = conn.execute(
                "SELECT table_id, table_name FROM _table_registry WHERE source_path = ?",
                (source_path,),
            ).fetchall()
            ids: list[str] = []
            for row in rows:
                conn.execute(f'DROP TABLE IF EXISTS "{row["table_name"]}"')
                ids.append(row["table_id"])
            conn.execute(
                "DELETE FROM _table_registry WHERE source_path = ?",
                (source_path,),
            )
            conn.commit()
            if ids:
                logger.info(
                    "Deleted %d table(s) for source '%s'", len(ids), source_path
                )
            return ids
        finally:
            conn.close()

    # ── Schema introspection ─────────────────────────────────

    def get_table_schemas(self, table_ids: list[str]) -> list[dict]:
        """Return schema info needed for SQL generation."""
        conn = self._get_conn()
        try:
            out: list[dict] = []
            for tid in table_ids:
                reg = conn.execute(
                    "SELECT * FROM _table_registry WHERE table_id = ?", (tid,)
                ).fetchone()
                if not reg:
                    continue
                cols = conn.execute(
                    f'PRAGMA table_info("{reg["table_name"]}")'
                ).fetchall()
                out.append(
                    {
                        "table_id": reg["table_id"],
                        "table_name": reg["table_name"],
                        "source_path": reg["source_path"],
                        "sheet_name": reg["sheet_name"],
                        "description": reg["description"],
                        "columns": [{"name": c["name"], "type": c["type"]} for c in cols],
                        "row_count": reg["row_count"],
                    }
                )
            return out
        finally:
            conn.close()

    # ── Ollama: SQL generation ───────────────────────────────

    async def generate_sql(
        self,
        user_query: str,
        table_schemas: list[dict],
    ) -> dict:
        """Generate a SQLite ``SELECT`` query from a natural-language question."""
        schema_parts: list[str] = []
        for s in table_schemas:
            cols = ", ".join(f'{c["name"]} ({c["type"]})' for c in s["columns"])
            schema_parts.append(
                f'Table "{s["table_name"]}": {s["description"]}\n'
                f"  Columns: {cols}  — {s['row_count']} rows"
            )

        prompt = (
            "You are a SQLite expert. Generate a SQL query to answer the user's question.\n\n"
            f"Available tables:\n{''.join(schema_parts)}\n\n"
            f"User question: {user_query}\n\n"
            "Respond with ONLY this JSON (no markdown):\n"
            '{"sql": "SELECT ...", "explanation": "brief explanation"}\n\n'
            "Rules:\n"
            "- ONLY SELECT statements.\n"
            "- Table and column names must match exactly.\n"
            "- Use SQLite syntax.\n"
            "- LIMIT results to 50 rows unless the user asks for all.\n"
            "- Use double-quotes for identifiers with special characters.\n"
        )

        return await self._ollama_json(prompt, self.text_to_sql_model)

    async def generate_sql_with_error_feedback(
        self,
        user_query: str,
        table_schemas: list[dict],
        previous_sql: str,
        error: str,
    ) -> dict:
        """Retry SQL generation with error context."""
        schema_parts: list[str] = []
        for s in table_schemas:
            cols = ", ".join(f'{c["name"]} ({c["type"]})' for c in s["columns"])
            schema_parts.append(
                f'Table "{s["table_name"]}": {s["description"]}\n'
                f"  Columns: {cols}"
            )

        prompt = (
            "You are a SQLite expert. Your previous query had an error. Fix it.\n\n"
            f"Available tables:\n{''.join(schema_parts)}\n\n"
            f"User question: {user_query}\n"
            f"Previous SQL: {previous_sql}\n"
            f"Error: {error}\n\n"
            "Respond with ONLY this JSON:\n"
            '{"sql": "SELECT ...", "explanation": "brief explanation"}\n'
        )

        return await self._ollama_json(prompt)

    async def _ollama_json(self, prompt: str, model: str | None = None) -> dict:
        """Send a prompt to Ollama and parse the JSON response."""
        resp = await self.http_client.post(
            f"{self.ollama_url}/api/generate",
            json={
                "model": model or self.ollama_model,
                "prompt": prompt,
                "stream": False,
                "format": "json",
                "options": {"temperature": 0.1, "num_predict": 500, "num_ctx": 4096},
                "keep_alive": "10m",
            },
        )
        resp.raise_for_status()
        raw = resp.json().get("response", "{}")
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            raise ValueError(f"Failed to parse Ollama JSON response: {raw[:200]}")

    # ── SQL execution ────────────────────────────────────────

    _FORBIDDEN_KEYWORDS = re.compile(
        r"\b(DROP|DELETE|INSERT|UPDATE|ALTER|CREATE|ATTACH|DETACH|REPLACE|VACUUM|REINDEX|PRAGMA|LOAD_EXTENSION)\b",
        re.IGNORECASE,
    )

    _FORBIDDEN_IDENTIFIERS = re.compile(
        r"\b(sqlite_master|sqlite_schema|sqlite_temp_master|sqlite_temp_schema|_table_registry)\b",
        re.IGNORECASE,
    )

    _MAX_RESULT_ROWS = 200

    def _validate_sql(self, sql: str) -> str:
        """Validate and sanitise an LLM-generated SQL string.

        Returns the cleaned single statement or raises ``ValueError``.
        """
        # Strip leading/trailing whitespace and trailing semicolons
        stripped = sql.strip().rstrip(";")

        # Block multiple statements (semicolons anywhere in the query body)
        if ";" in stripped:
            raise ValueError(
                "Multiple SQL statements are not allowed"
            )

        if not stripped.upper().startswith("SELECT"):
            raise ValueError("Only SELECT queries are allowed")

        if self._FORBIDDEN_KEYWORDS.search(stripped):
            raise ValueError("Query contains forbidden SQL keywords")

        if self._FORBIDDEN_IDENTIFIERS.search(stripped):
            raise ValueError(
                "Query references internal/system tables which is not allowed"
            )

        # Validate that only known user tables are referenced
        known = self._get_known_table_names()
        referenced = self._extract_table_refs(stripped)
        unknown = referenced - known
        if unknown:
            raise ValueError(
                f"Query references unknown tables: {', '.join(sorted(unknown))}"
            )

        return stripped

    def _get_known_table_names(self) -> set[str]:
        """Return the set of user-created table names from the registry."""
        conn = self._get_conn(readonly=True)
        try:
            rows = conn.execute(
                "SELECT table_name FROM _table_registry"
            ).fetchall()
            return {r["table_name"] for r in rows}
        finally:
            conn.close()

    @staticmethod
    def _extract_table_refs(sql: str) -> set[str]:
        """Best-effort extraction of table names from a SELECT statement.

        Looks for identifiers after FROM and JOIN keywords.
        """
        # Remove string literals to avoid false matches
        cleaned = re.sub(r"'[^']*'", "''", sql)
        pattern = re.compile(
            r"(?:FROM|JOIN)\s+[\"\`]?([a-zA-Z_][a-zA-Z0-9_]*)[\"\`]?",
            re.IGNORECASE,
        )
        return {m.group(1).lower() for m in pattern.finditer(cleaned)}

    def execute_sql(self, sql: str) -> dict:
        """Execute a validated, read-only SQL query.

        Opens the database in **read-only mode** so that even if
        validation is bypassed, no writes can occur.

        Raises ``ValueError`` on forbidden statements or execution errors.
        """
        validated = self._validate_sql(sql)

        # Enforce a row limit to prevent memory exhaustion
        if not re.search(r"\bLIMIT\b", validated, re.IGNORECASE):
            validated = f"{validated} LIMIT {self._MAX_RESULT_ROWS}"

        conn = self._get_conn(readonly=True)
        try:
            cursor = conn.execute(validated)
            columns = (
                [desc[0] for desc in cursor.description] if cursor.description else []
            )
            rows = [dict(r) for r in cursor.fetchmany(self._MAX_RESULT_ROWS)]
            return {"columns": columns, "rows": rows, "row_count": len(rows)}
        except sqlite3.OperationalError as exc:
            raise ValueError(f"SQL execution error: {exc}")
        finally:
            conn.close()

    # ── Cleanup ──────────────────────────────────────────────

    async def close(self) -> None:
        if self._http_client is not None:
            await self._http_client.aclose()
            self._http_client = None

    # ── Internal helpers ─────────────────────────────────────

    @staticmethod
    def _coerce_value(val: Any, expected_type: str) -> Any:
        if val is None or (isinstance(val, str) and val.strip().lower() in ("", "null", "none", "nan")):
            return None
        try:
            if expected_type == "INTEGER":
                return int(float(str(val)))
            if expected_type in ("REAL", "NUMERIC"):
                return float(str(val))
            if expected_type == "BOOLEAN":
                return 1 if str(val).lower() in ("true", "1", "yes", "ja") else 0
            return str(val)
        except (ValueError, TypeError):
            return str(val)
